#!/usr/bin/env python3
"""
Fetch all Aquatic Exotica products and export Junglyst bulk-upload files.

Usage:
    python scripts/export_aquatic_exotica_catalog.py

Outputs (under imports/seller_onboarding/):
    - aquatic_exotica_bulk_upload.xlsx
    - junglyst_seller_product_import_template.csv
"""
from __future__ import annotations

import csv
import json
import re
import urllib.request
from decimal import Decimal, ROUND_HALF_UP
from html import unescape
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

API_BASE = "https://api.aquaticexotica.com/api/products/"
BRAND_NAME = "Aquatic Exotica"
SELLER_SLUG_HINT = "aquaticexotica"
DEFAULT_GST_RATE = Decimal("12")
DEFAULT_COMMISSION_RATE = Decimal("10")
DEFAULT_VARIANT_NAME = "Standard"

# Best-effort mapping from Aquatic Exotica category names → Junglyst seed_categories tree.
CATEGORY_MAP = {
    "Aquatic Plants": ("Plants", "Aquatic Plants"),
    "Rhizome plants": ("Plants", "Aquatic Plants"),
    "Moss": ("Terrarium & Paludarium", "Terrarium Moss"),
    "Terrarium Plants": ("Terrarium & Paludarium", "Terrarium Plants"),
    "Indoor Plants": ("Terrarium & Paludarium", "Terrarium Plants"),
    "Exotic Plants": ("Terrarium & Paludarium", "Terrarium Plants"),
    "Premium": ("Plants", "Rare & Exotic"),
}

BULK_COLUMNS = [
    "brand_name",
    "seller_store_name",
    "external_product_id",
    "name",
    "tagline",
    "description",
    "scientific_name",
    "origin",
    "care_level",
    "light_requirements",
    "growth_rate",
    "co2_requirement",
    "water_temperature",
    "ph_range",
    "is_rare",
    "is_active",
    "junglyst_category",
    "junglyst_subcategory",
    "source_categories",
    "source_tags",
    "variant_name",
    "base_price",
    "compare_at_price",
    "gst_rate",
    "commission_rate",
    "stock",
    "weight_kg",
    "length_cm",
    "width_cm",
    "height_cm",
    "item_category",
    "packed_weight_grams",
    "image_url_1",
    "image_url_2",
    "image_url_3",
    "source_retail_price",
    "source_compare_at_price",
    "source_discount_percent",
    "source_rating",
    "source_is_in_stock",
    "import_notes",
]

TEMPLATE_SAMPLE_ROW = {
    "brand_name": "Your Brand",
    "seller_store_name": "your-store-slug",
    "external_product_id": "SKU-001",
    "name": "Anubias nana Petite (1 Pot)",
    "tagline": "Compact rhizome plant for nano tanks",
    "description": "Short plain-text description for the listing.",
    "scientific_name": "Anubias barteri var. nana 'Petite'",
    "origin": "Cultivated",
    "care_level": "Easy",
    "light_requirements": "Low",
    "growth_rate": "Slow",
    "co2_requirement": "Low",
    "water_temperature": "22-28 C",
    "ph_range": "6.0-7.5",
    "is_rare": "FALSE",
    "is_active": "TRUE",
    "junglyst_category": "Aquascaping",
    "junglyst_subcategory": "Starter Packs",
    "source_categories": "Aquatic Plants|Rhizome plants",
    "source_tags": "Rhizomes|Aquatic Plants",
    "variant_name": "Standard",
    "base_price": "262.30",
    "compare_at_price": "450.00",
    "gst_rate": "12",
    "commission_rate": "10",
    "stock": "5",
    "weight_kg": "0.5",
    "length_cm": "10",
    "width_cm": "10",
    "height_cm": "10",
    "item_category": "light",
    "packed_weight_grams": "200",
    "image_url_1": "https://example.com/image.png",
    "image_url_2": "",
    "image_url_3": "",
    "source_retail_price": "320.00",
    "source_compare_at_price": "450.00",
    "source_discount_percent": "28",
    "source_rating": "5.0",
    "source_is_in_stock": "TRUE",
    "import_notes": "Optional notes for your team",
}


def fetch_all_products() -> list[dict]:
    page = 1
    results: list[dict] = []
    while True:
        url = f"{API_BASE}?page={page}"
        with urllib.request.urlopen(url, timeout=60) as response:
            data = json.loads(response.read().decode())
        results.extend(data.get("results") or [])
        if not data.get("next"):
            break
        page += 1
    return results


def strip_html(html: str, max_len: int | None = None) -> str:
    if not html:
        return ""
    text = re.sub(r"<[^>]+>", " ", html)
    text = unescape(re.sub(r"\s+", " ", text)).strip()
    if max_len and len(text) > max_len:
        return text[: max_len - 3] + "..."
    return text


def money(value) -> str:
    if value in (None, ""):
        return ""
    return str(Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def calc_base_price(retail_price, gst_rate: Decimal, commission_rate: Decimal) -> str:
    if retail_price in (None, ""):
        return ""
    retail = Decimal(str(retail_price))
    divisor = Decimal("1") + (gst_rate / Decimal("100")) + (commission_rate / Decimal("100"))
    if divisor == 0:
        return ""
    base = (retail / divisor).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return str(base)


def pick_category_names(product: dict) -> tuple[str, str, str]:
    names = [c.get("name", "") for c in (product.get("categories") or []) if c.get("name")]
    names_set = set(names)
    source = " | ".join(names)
    
    mapped_subs = []
    
    # 1. Aquatic Mosses (e.g. Christmas Moss)
    # Aquatic Mosses (e.g. Christmas Moss) → Plants > Aquatic Moss + Terrarium moss + Aquatic plants +  Terrarium & Paludarium
    if "Moss" in names_set and "Aquatic Plants" in names_set:
        mapped_subs.append(("Plants", "Aquatic Moss"))
        mapped_subs.append(("Terrarium & Paludarium", "Terrarium Moss"))
        mapped_subs.append(("Plants", "Aquatic Plants"))

    # 2. Terrarium Mosses (e.g. Sheet/Mood Moss)
    # Terrarium Mosses (e.g. Sheet/Mood Moss) → Terrarium & Paludarium > Terrarium moss + Plants
    elif "Moss" in names_set:
        mapped_subs.append(("Terrarium & Paludarium", "Terrarium Moss"))
        mapped_subs.append(("Plants", "Mosses"))

    # 3. Rare/Exotic Rhizomes & Aquatic Plants (e.g. Bucephalandra)
    # Rare/Exotic Rhizomes & Aquatic Plants (e.g. Bucephalandra) → Plants > Rare & Exotic + Aquatic plants + Terrarium & Paludarium
    elif ("Rhizome plants" in names_set or "Aquatic Plants" in names_set) and ("Exotic Plants" in names_set or "Premium" in names_set):
        mapped_subs.append(("Plants", "Rare & Exotic"))
        mapped_subs.append(("Plants", "Aquatic Plants"))
        mapped_subs.append(("Terrarium & Paludarium", "Terrarium Plants"))

    # 4. Standard Rhizomes & Aquatic Plants (e.g. Anubias)
    # Standard Rhizomes & Aquatic Plants (e.g. Anubias) → Plants > Aquatic Plants + Terrarium & Paludarium
    elif "Rhizome plants" in names_set or "Aquatic Plants" in names_set:
        mapped_subs.append(("Plants", "Aquatic Plants"))
        mapped_subs.append(("Terrarium & Paludarium", "Terrarium Plants"))

    # 5. Exotic/Indoor/Terrarium Plants (Non-Aquatic) (e.g. Peperomia, Fittonia, Ferns)
    # Exotic/Indoor/Terrarium Plants (Non-Aquatic) (e.g. Peperomia, Fittonia, Ferns) → Terrarium & Paludarium > Terrarium Plants
    elif "Exotic Plants" in names_set or "Indoor Plants" in names_set or "Terrarium Plants" in names_set:
        mapped_subs.append(("Terrarium & Paludarium", "Terrarium Plants"))

    # Deduplicate mapped_subs
    mapped_subs = list(dict.fromkeys(mapped_subs))

    # Fallback to single category map
    if not mapped_subs:
        for name in names:
            if name in CATEGORY_MAP:
                mapped_subs.append(CATEGORY_MAP[name])
                break

    if not mapped_subs:
        if names:
            return "REVIEW_NEEDED", "", source
        return "", "", source

    cats = list(dict.fromkeys([c for c, _ in mapped_subs]))
    subs = list(dict.fromkeys([s for _, s in mapped_subs]))

    return "|".join(cats), "|".join(subs), source


def pick_tags(product: dict) -> str:
    tags = product.get("tagDetails") or []
    return " | ".join(t.get("name", "") for t in tags if t.get("name"))


def infer_care_from_description(description_html: str) -> dict[str, str]:
    text = strip_html(description_html).lower()
    care = "Easy"
    if "advanced" in text:
        care = "Advanced"
    elif "medium" in text and "care" in text:
        care = "Medium"

    light = "Medium"
    if "low light" in text or "low to medium" in text or "low lighting" in text:
        light = "Low"
    elif "high light" in text or "high lighting" in text:
        light = "High"

    growth = "Moderate"
    if "slow growth" in text or "extremely slow" in text or "slow-growing" in text:
        growth = "Slow"
    elif "fast growth" in text or "fast-growing" in text:
        growth = "Fast"

    co2 = "Low"
    if "co2 required" in text or "co2 injection" in text:
        co2 = "High"
    elif "co2 recommended" in text or "medium co2" in text:
        co2 = "Medium"

    return {
        "care_level": care,
        "light_requirements": light,
        "growth_rate": growth,
        "co2_requirement": co2,
    }


def product_to_row(product: dict) -> dict[str, str]:
    cat, sub, source_categories = pick_category_names(product)
    care = infer_care_from_description(product.get("description") or "")
    gst = DEFAULT_GST_RATE
    commission = DEFAULT_COMMISSION_RATE
    retail = product.get("price")
    base_price = calc_base_price(retail, gst, commission)

    notes = []
    if cat == "REVIEW_NEEDED":
        notes.append("Map junglyst_category/subcategory before import.")
    if not product.get("isInStock"):
        notes.append("Out of stock at source.")

    return {
        "brand_name": BRAND_NAME,
        "seller_store_name": SELLER_SLUG_HINT,
        "external_product_id": str(product.get("id", "")),
        "name": product.get("name") or "",
        "tagline": "",
        "description": strip_html(product.get("description") or "", max_len=8000),
        "scientific_name": "",
        "origin": "",
        "care_level": care["care_level"],
        "light_requirements": care["light_requirements"],
        "growth_rate": care["growth_rate"],
        "co2_requirement": care["co2_requirement"],
        "water_temperature": "",
        "ph_range": "",
        "is_rare": "FALSE",
        "is_active": "TRUE" if product.get("isActive", True) else "FALSE",
        "junglyst_category": cat if cat != "REVIEW_NEEDED" else "",
        "junglyst_subcategory": sub,
        "source_categories": source_categories,
        "source_tags": pick_tags(product),
        "variant_name": DEFAULT_VARIANT_NAME,
        "base_price": base_price,
        "compare_at_price": money(product.get("compareAtPrice")),
        "gst_rate": str(gst),
        "commission_rate": str(commission),
        "stock": str(product.get("stock") or 0),
        "weight_kg": "0.5",
        "length_cm": "10",
        "width_cm": "10",
        "height_cm": "10",
        "item_category": "light",
        "packed_weight_grams": "200",
        "image_url_1": product.get("imageUrl") or "",
        "image_url_2": product.get("thumbnailUrl") or "",
        "image_url_3": "",
        "source_retail_price": money(retail),
        "source_compare_at_price": money(product.get("compareAtPrice")),
        "source_discount_percent": str(product.get("discountPercentage") or ""),
        "source_rating": str(product.get("rating") or ""),
        "source_is_in_stock": "TRUE" if product.get("isInStock") else "FALSE",
        "import_notes": "; ".join(notes),
    }


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=BULK_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def autosize_columns(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def write_workbook(
    path: Path,
    product_rows: list[dict[str, str]],
) -> None:
    wb = Workbook()

    # Products sheet
    ws = wb.active
    ws.title = "Products"
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(BULK_COLUMNS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in product_rows:
        ws.append([row.get(col, "") for col in BULK_COLUMNS])
    ws.freeze_panes = "A2"
    autosize_columns(ws)

    # Instructions sheet
    inst = wb.create_sheet("Instructions")
    instructions = [
        ["Junglyst seller product bulk import"],
        [""],
        [f"Brand: {BRAND_NAME}"],
        [f"Source API: {API_BASE}"],
        [""],
        ["How to use this file"],
        ["1. Review junglyst_category and junglyst_subcategory columns — adjust to match your live Junglyst category tree."],
        ["2. base_price is estimated from source_retail_price using gst_rate and commission_rate (buyer price formula)."],
        ["3. Update seller_store_name to the Aquatic Exotica seller slug after the seller account is created."],
        ["4. Fill weight/dimensions per SKU before go-live."],
        ["5. Import via admin-create API or future bulk importer using the same column names."],
        [""],
        ["Required for create"],
        ["name, seller_id (or seller_store_name lookup), description, variants.base_price, variants.stock"],
        [""],
        ["Allowed enum values"],
        ["care_level: Easy | Medium | Advanced"],
        ["light_requirements: Low | Medium | High"],
        ["growth_rate: Slow | Moderate | Fast"],
        ["co2_requirement: Low | Medium | High"],
        ["item_category: light | heavy"],
        ["is_rare / is_active: TRUE | FALSE"],
        [""],
        ["Pricing formula"],
        ["buyer_price = base_price * (1 + gst_rate/100 + commission_rate/100)"],
    ]
    for line in instructions:
        inst.append(line)
    inst.column_dimensions["A"].width = 100

    # Category reference
    ref = wb.create_sheet("CategoryMapping")
    ref.append(["source_category", "suggested_junglyst_category", "suggested_junglyst_subcategory"])
    for source, (jung_cat, jung_sub) in sorted(CATEGORY_MAP.items()):
        ref.append([source, jung_cat, jung_sub])
    autosize_columns(ref)

    # Raw API snapshot (for audit)
    raw = wb.create_sheet("SourceAPI")
    raw.append(["Fetched from Aquatic Exotica API — keep for traceability"])
    raw.append([])

    wb.save(path)


def main() -> None:
    out_dir = Path(__file__).resolve().parent.parent / "imports" / "seller_onboarding"
    out_dir.mkdir(parents=True, exist_ok=True)

    products = fetch_all_products()
    rows = [product_to_row(p) for p in products]

    xlsx_path = out_dir / "aquatic_exotica_bulk_upload.xlsx"
    csv_data_path = out_dir / "aquatic_exotica_bulk_upload.csv"
    template_path = out_dir / "junglyst_seller_product_import_template.csv"

    write_workbook(xlsx_path, rows)
    write_csv(csv_data_path, rows)
    write_csv(template_path, [TEMPLATE_SAMPLE_ROW])

    print(f"Fetched {len(products)} products")
    print(f"Wrote {xlsx_path}")
    print(f"Wrote {csv_data_path}")
    print(f"Wrote {template_path}")


if __name__ == "__main__":
    main()
